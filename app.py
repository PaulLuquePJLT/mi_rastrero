# app.py
# =============================================================
#       Rastrero (Hasbro) – versión Streamlit
# =============================================================
import streamlit as st
import pandas as pd
import numpy as np
from io import BytesIO
import re, gc, datetime as dt
import openpyxl
from streamlit_option_menu import option_menu

# -------------------------------------------------------------
# 1. Configuración global y estilos
# -------------------------------------------------------------
st.set_page_config(page_title="Generador de Rastrero Hasbro",
                   page_icon="📦",
                   layout="wide")

FA_LINK = """
<link rel="stylesheet"
      href="https://cdnjs.cloudflare.com/ajax/libs/font-awesome/6.5.0/css/all.min.css">
"""
st.markdown(FA_LINK, unsafe_allow_html=True)

CUSTOM_CSS = """
<style>
h1, h2, h3, h4 { font-family:'Inter', sans-serif; }
.header {
    display:flex; align-items:center; gap:18px;
    background:linear-gradient(135deg,#0d47a1 0%,#1976d2 100%);
    color:#fff; padding:18px 26px; border-radius:8px;
}
.header img { width:70px; height:70px; object-fit:contain;
              background:#fff; border-radius:50%;
              padding:6px; box-shadow:0 2px 6px rgba(0,0,0,.2); }
.status-ok   { color:#2e7d32; }
.status-warn { color:#c62828; }
</style>
"""
st.markdown(CUSTOM_CSS, unsafe_allow_html=True)

# -------------------------------------------------------------
# 2. Cabecera y barra de estado
# -------------------------------------------------------------
with st.container():
    col1, col2 = st.columns([1, 10])
    with col1:
        st.image(
            "https://images.seeklogo.com/logo-png/24/1/hasbro-logo-png_seeklogo-241282.png",
            width=120
        )
    with col2:
        st.markdown("""
        <h2 style="margin-bottom:0;">Generador de Rastrero Hasbro</h2>
        <small>Developed by: <strong>PJLT</strong></small>
        """, unsafe_allow_html=True)

status_placeholder = st.empty()
progress_bar       = st.progress(0)

def update_status(msg: str, pct: int|None = None, ok: bool = True):
    icon = "check-circle" if ok else "exclamation-triangle"
    clazz = "status-ok" if ok else "status-warn"
    status_placeholder.markdown(
        f"<span class='{clazz}'><i class='fa fa-{icon}'></i> {msg}</span>",
        unsafe_allow_html=True
    )
    if pct is not None:
        progress_bar.progress(pct)

# -------------------------------------------------------------
# 3. Utilidades comunes
# -------------------------------------------------------------
def clean_lote(series: pd.Series) -> pd.Series:
    return (series.astype(str)
                  .str.replace(r'\s+', '', regex=True)
                  .str.replace('\u00A0', '')
                  .str.strip())


def clean_number(col: pd.Series) -> pd.Series:
    s = (col.astype(str)
             .str.replace(r'[^0-9,.-]', '', regex=True)
             .str.replace('.', '', regex=False)
             .str.replace(',', '.', regex=False))
    return pd.to_numeric(s, errors='coerce').fillna(0.0)


def preparar_stock(file_bytes: bytes) -> pd.DataFrame:
    df = pd.read_excel(BytesIO(file_bytes))
    df.columns = (df.columns.str.normalize('NFKD')
                            .str.encode('ascii','ignore').str.decode('utf-8')
                            .str.strip())
    df['Lote Proveedor']  = clean_lote(df['Lote Proveedor'])
    df['Cant. Final UMS'] = clean_number(df['Cant. Final UMS'])
    df['Concat_U_A']      = df['Ubicacion'] + df['Cod. Articulo']

    def factor(s):
        n = s.astype(str).str.extract(r'C(\d+(?:\.\d+)?)U', expand=False)
        return pd.to_numeric(n, errors='coerce').fillna(1)
    df['Factor'] = factor(df['Huella'])
    df['Cajas']  = df['Cant. Final UMS'] / df['Factor']
    df['UM']     = 'CAJ'

    def nivel(s):
        ult = s.str[-1]
        return np.where(ult.str.isnumeric() & (ult.astype(int) < 3), 'BAJO', 'ALTO')
    df['Nivel'] = nivel(df['Ubicacion'])

    keep = ['Concat_U_A','Ubicacion','Cod. Articulo','Factor','UM',
            'Nivel','Lote Proveedor','Cant. Final UMS','Cajas']
    grp  = ['Concat_U_A','Ubicacion','Cod. Articulo',
            'Factor','UM','Nivel','Lote Proveedor']
    df_bd = (df[keep].groupby(grp, as_index=False)
                     .agg({'Cant. Final UMS':'sum','Cajas':'sum'})
                     .rename(columns={'Ubicacion':'Ubicación',
                                      'Cod. Articulo':'Cod. Artículo'}))
    return df_bd

# Helpers para Rastrero Out

def norm_cols(df: pd.DataFrame) -> pd.DataFrame:
    df = df.copy()
    df.columns = (df.columns.str.normalize('NFKD')
                              .str.encode('ascii','ignore').str.decode('utf-8')
                              .str.strip())
    return df


def factor(s: pd.Series) -> pd.Series:
    return (s.astype(str)
             .str.extract(r'C(\d+(?:\.\d+)?)U', expand=False)
             .astype(float)
             .fillna(1))


def calc_pasillo(u: str) -> str:
    if len(u) < 11:
        return 'Libre'
    if u[3:5] == 'MR':
        return 'Pasillo_1'
    tramo = u[8:11]
    if tramo in ('C06','C07','C08'):
        return 'Pasillo_1'
    if tramo in ('C09','C10'):
        return 'Pasillo_2'
    if tramo in ('C11','C12'):
        return 'Pasillo_3'
    return 'Libre'

# -------------------------------------------------------------
# 4. Estado global en sesión
# -------------------------------------------------------------
if 'state_in' not in st.session_state:
    st.session_state.state_in = {}
if 'state_out' not in st.session_state:
    st.session_state.state_out = {}

# =============================================================
#   5. MÓDULO – RASTRERO IN (con resumen dinámico tras filtros)
# =============================================================

def render_summary_block(n_regs, sum_cajas, rango, value_font_size='24px', date_font_size='20px'):
    """
    Renderiza un bloque de resumen con tres tarjetas: Registros, Cajas y Fechas.
    """
    icon_map = ['list', 'cube', 'calendar']
    labels   = ['Registros', 'Cajas', 'Fechas']
    values   = [n_regs, sum_cajas, rango]
    colors   = ['#42a5f5', '#66bb6a', '#ab47bc']

    cols = st.columns(3, gap='large')
    for col, icon, label, val, bg in zip(cols, icon_map, labels, values, colors):
        fs = date_font_size if (label == 'Fechas' and date_font_size) else value_font_size
        col.markdown(f"""
            <div style="
                background:{bg}; color:white; border-radius:8px;
                padding:20px; display:flex; align-items:center;
                min-height:100px;"
            >
              <i class='fa fa-{icon}' style='font-size:32px; margin-right:16px;'></i>
              <div>
                <div style='font-size:{fs}; font-weight:bold;'>{val}</div>
                <div style='font-size:12px; opacity:0.8; text-transform:uppercase;'>{label}</div>
              </div>
            </div>
        """, unsafe_allow_html=True)


def rastrero_in():
    st.subheader("📥 Rastrero In")
    state = st.session_state.setdefault('state_in', {})

    # --- Carga de archivos ---
    with st.container():
        st.markdown("### 📁 Carga de archivos")
        u1, u2, u3 = st.columns([1,1,1], gap='medium')
        with u1:
            mov_file = st.file_uploader("Movimientos internos", type="xlsx", key="mov_int")
        with u2:
            stock_file = st.file_uploader("Stock", type="xlsx", key="stock_in")
        with u3:
            tmpl_file = st.file_uploader("Plantilla", type="xlsx", key="tmpl_in")
    st.markdown("<hr style='margin:8px 0'>", unsafe_allow_html=True)

    # Guardar bytes de plantilla para exportación (Rastrero In)
    if tmpl_file:
        try:
            state['tmpl_in_bytes'] = tmpl_file.read()
        except Exception:
            pass

    # --- Procesar Movimientos Internos ---
    if mov_file and 'T_Mov_Internos' not in state:
        name = mov_file.name
        if not name.startswith("ReportConsultasMovimientosInternos"):
            update_status("⚠️ Archivo Movimientos Internos incorrecto", ok=False)
        else:
            update_status("Leyendo Movimientos Internos…", 10)
            df = pd.read_excel(mov_file)
            invalid_refs = ["PICKING-TRASLADO","PICKING-SURTIDO","PICKING-PTS-SURTIDO","LPN-TRASLADO"]
            df = df[~df['Referencia 1'].isin(invalid_refs)]
            df = df[df['UM Origen'] != 'UNIDAD']
            df['Glosa'] = df['Glosa'].fillna('Movimiento desde DWEB')
            df['Categoria'] = ''
            mask_alma = (df['Motivo']=='CAMBIO DE UBICACION') & df['Ubicación Origen'].str.startswith('B4.RE', na=False)
            mask_mov  = (df['Motivo']=='CAMBIO DE UBICACION') & ~mask_alma
            mask_est  = df['Motivo']=='CAMBIO DE ESTADO'
            df.loc[mask_alma, 'Categoria'] = 'Almacenamiento'
            df.loc[mask_mov,  'Categoria'] = 'Mov_Interno'
            df.loc[mask_est,  'Categoria'] = 'Regul_Interna/Cambio_Estado'
            df['Concat_MI'] = df['Ubicación Destino'].astype(str) + df['Cod. Articulo'].astype(str)
            df['Fecha Movimiento'] = pd.to_datetime(df['Fecha Movimiento'])
            state['T_Mov_Internos'] = df
            update_status("Movimientos Internos listos ✓", 30)

    # --- Procesar Stock para T_Stock_I ---
    if stock_file and 'T_Stock_I' not in state:
        update_status("Leyendo Stock Interno…", 40)
        df_s = pd.read_excel(stock_file)
        df_s = norm_cols(df_s)
        df_s['Ubicación_1'] = df_s['Ubicacion']
        df_s['Cod. Artículo_1'] = df_s['Cod. Articulo']
        df_s['UM_1'] = df_s['UM']
        df_stock_i = (df_s.groupby(
            ['Ubicación_1','Cod. Artículo_1','UM_1'],
            as_index=False
        ).agg({'Cant. Final': 'sum'})
                      .rename(columns={'Cant. Final':'Stock Final_1'}))
        df_stock_i['Concat_ST'] = (
            df_stock_i['Ubicación_1'].astype(str)
            + df_stock_i['Cod. Artículo_1'].astype(str)
        )
        state['T_Stock_I'] = df_stock_i
        update_status("Stock Interno listo ✓", 50)

    # --- Filtros y preparación de df filtrado ---
    df_f = pd.DataFrame()
    if 'T_Mov_Internos' in state:
        df_all = state['T_Mov_Internos']
        # Inicialización de opciones y estado
        all_cats   = sorted(df_all['Categoria'].dropna().unique())
        all_glosas = sorted(df_all['Glosa'].dropna().unique())
        all_lotes  = sorted(df_all['Lote Proveedor Destino'].astype(str).unique())
        st.session_state.setdefault('cat_sel', all_cats.copy())
        st.session_state.setdefault('glosa_sel', all_glosas.copy())
        st.session_state.setdefault('lote_sel', all_lotes.copy())

        with st.expander("🎛️ Filtros", expanded=True):
            c1, c2 = st.columns(2, gap='large')
            with c1:
                st.multiselect("Categoría", all_cats, key='cat_sel')
            with c2:
                glosa_opts = sorted(
                    df_all[df_all['Categoria'].isin(st.session_state.cat_sel)]['Glosa']
                    .dropna().unique()
                )
                st.session_state['glosa_sel'] = [g for g in st.session_state.glosa_sel if g in glosa_opts] or glosa_opts
                st.multiselect("Glosa", glosa_opts, key='glosa_sel')
            lote_opts = sorted(
                df_all[(df_all['Categoria'].isin(st.session_state.cat_sel)) &
                       (df_all['Glosa'].isin(st.session_state.glosa_sel))]
                ['Lote Proveedor Destino'].astype(str).unique()
            )
            st.session_state['lote_sel'] = [l for l in st.session_state.lote_sel if l in lote_opts] or lote_opts
            st.multiselect("Lote", lote_opts, key='lote_sel')

        # Construir df filtrado con los valores actuales seleccionados
        df_f = df_all[
            df_all['Categoria'].isin(st.session_state.cat_sel) &
            df_all['Glosa'].isin(st.session_state.glosa_sel) &
            df_all['Lote Proveedor Destino'].astype(str).isin(st.session_state.lote_sel)
        ]
    # --- Bloques de resumen (sobre df filtrado) ---
    if not df_f.empty:
        n_regs_f    = len(df_f)
        sum_cajas_f = df_f['Cant. Destino'].sum()
        try:
            min_d_f = df_f['Fecha Movimiento'].min()
            max_d_f = df_f['Fecha Movimiento'].max()
            dias  = ['Lun','Mar','Mié','Jue','Vie','Sáb','Dom']
            meses = ['Ene','Feb','Mar','Abr','May','Jun','Jul','Ago','Sep','Oct','Nov','Dic']
            def _fmt(d):
                return f"{dias[d.weekday()]} {d.day} de {meses[d.month-1]}"
            rango_f = _fmt(min_d_f) if min_d_f == max_d_f else f"Desde {_fmt(min_d_f)} hasta {_fmt(max_d_f)}"
        except Exception:
            rango_f = ""
        st.markdown("<hr style='margin:8px 0'>", unsafe_allow_html=True)
        render_summary_block(n_regs_f, sum_cajas_f, rango_f, value_font_size='24px', date_font_size='20px')

    # Mostrar controles de generación y exportación (Rastrero In)
    if 'T_Stock_I' in state and 'T_Mov_Internos' in state:
        st.markdown("### 🧰 Controles")
        col_right, col_left = st.columns([2,1])
        with col_right:
            b1, b2 = st.columns([1,1])
            gen_clicked = b1.button("⚙️ Generar Rastrero In", use_container_width=True)
            # Exportación: requiere plantilla, tablas generadas y fecha seleccionada
            fecha_val = st.session_state.get('fecha_in', None)
            export_ready = ('tmpl_in_bytes' in state and 'R_Nivel_Bajo' in state and 'R_Nivel_Alto' in state and isinstance(fecha_val, dt.date))
            if export_ready:
                try:
                    _wb = openpyxl.load_workbook(BytesIO(state['tmpl_in_bytes']))
                    faltantes = [h for h in ("R_Nivel_Bajo","R_Nivel_Alto") if h not in _wb.sheetnames]
                    if faltantes:
                        b2.button("⬇️ Exportar Excel Ingresos", disabled=True, use_container_width=True)
                        st.error(f"No se encontraron las hojas: {', '.join(faltantes)} en la plantilla.")
                    else:
                        def make_xlsx_in():
                            wb2 = openpyxl.load_workbook(BytesIO(state['tmpl_in_bytes']))
                            for hoja, dfp in (("R_Nivel_Bajo", state['R_Nivel_Bajo']), ("R_Nivel_Alto", state['R_Nivel_Alto'])):
                                ws = wb2[hoja]
                                # Pegar bloque desde C13
                                for i, row in enumerate(dfp.itertuples(index=False), start=13):
                                    for j, val in enumerate(row, start=3):  # C=3
                                        ws.cell(row=i, column=j, value=val)
                                # Fecha en I1
                                ws['I1'] = fecha_val.strftime('%d/%m/%Y')
                                # Lotes únicos en L1
                                lotes_df = state.get('T_Lotes_Filtrados')
                                if lotes_df is not None and not lotes_df.empty:
                                    for idx, lote in enumerate(lotes_df['Lote'], start=1):
                                        ws.cell(row=idx, column=12, value=lote)  # L=12
                                # Área de impresión B1:I{last_row}
                                last_row = 12 + len(dfp)
                                ws.print_area = f"B1:I{last_row}"
                            buf = BytesIO(); wb2.save(buf); buf.seek(0); return buf
                        fname = f"FORMATO_RASTRERO_INGRESOS_{fecha_val.strftime('%d.%m.%Y')}.xlsx"
                        b2.download_button(
                            "⬇️ Exportar Excel Ingresos",
                            data=make_xlsx_in(),
                            file_name=fname,
                            mime="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
                            use_container_width=True
                        )
                except Exception as e:
                    b2.button("⬇️ Exportar Excel Ingresos", disabled=True, use_container_width=True)
                    st.error(f"Error al leer la plantilla: {e}")
            else:
                b2.button("⬇️ Exportar Excel Ingresos", disabled=True, use_container_width=True)
        with col_left:
            # Picker sin fecha por defecto (None)
            st.date_input("📅 Fecha del reporte", value=st.session_state.get('fecha_in'), key='fecha_in')

        # Al hacer clic en Generar, construir tablas usando filtros
        if gen_clicked:
            if df_f.empty:
                st.error("No registros para filtros seleccionados")
            else:
                df_mov = df_f.copy()
                df_stock = state['T_Stock_I']
                # 1) Agrupar Mov (T_Cruce_In)
                cruce = df_mov.groupby('Concat_MI', as_index=False).agg({
                    'Ubicación Destino': 'first',
                    'Cod. Articulo': 'first',
                    'Cant. Destino': 'sum'
                }).rename(columns={
                    'Concat_MI':'Clave',
                    'Ubicación Destino':'Ubicacion Origen_1',
                    'Cod. Articulo':'Producto_1',
                    'Cant. Destino':'Ingresos_1'
                })
                cruce['UM_1'] = 'CJ'
                # 2) Stock por Clave
                stock_sum = state['T_Stock_I'].groupby('Concat_ST', as_index=False)['Stock Final_1'].sum()
                cruce = cruce.merge(stock_sum, left_on='Clave', right_on='Concat_ST', how='left').fillna({'Stock Final_1':0})
                # 3) Stock Inicial condicional + Observación
                ge_mask = cruce['Stock Final_1'] >= cruce['Ingresos_1']
                cruce['Stock Inicial_1'] = np.where(ge_mask, cruce['Stock Final_1'] - cruce['Ingresos_1'], cruce['Stock Final_1'])
                cruce['Observacion_1'] = np.where(ge_mask, '', 'Regu')
                # 4) Nivel_I por último dígito de Ubicacion Origen_1
                u = cruce['Ubicacion Origen_1'].astype(str).str.strip()
                last_char = u.str[-1]
                last_digit = last_char.where(last_char.str.isnumeric(), np.nan)
                cruce['Nivel_I'] = np.where(
                    last_digit.notna() & (last_digit.astype(int) <= 2), 'Bajo',
                    np.where(last_digit.notna() & (last_digit.astype(int) >= 3), 'Alto', '')
                )
                cruce['Check_1'] = cruce.get('Check_1', '') if isinstance(cruce.get('Check_1', ''), pd.Series) else ''
                state['T_Cruce_In'] = cruce

                # 5) Tablas finales por nivel
                cols_map = [
                    ('Ubicacion Origen','Ubicacion Origen_1'),
                    ('Producto','Producto_1'),
                    ('UM','UM_1'),
                    ('Stock Inicial','Stock Inicial_1'),
                    ('Ingresos','Ingresos_1'),
                    ('Stock Final','Stock Final_1'),
                    ('Check','Check_1'),
                    ('Observacion','Observacion_1')
                ]
                def build_r(df_src):
                    out = df_src[[src for (_,src) in cols_map]].copy()
                    out.columns = [dest for (dest,_) in cols_map]
                    return out
                r_bajo = build_r(cruce[cruce['Nivel_I'] == 'Bajo'])
                r_alto = build_r(cruce[cruce['Nivel_I'] == 'Alto'])
                state['R_Nivel_Bajo'] = r_bajo
                state['R_Nivel_Alto'] = r_alto

                # 6) Tabla de lotes filtrados
                lotes_df = pd.DataFrame({'Lote': sorted(df_mov['Lote Proveedor Destino'].astype(str).unique())})
                state['T_Lotes_Filtrados'] = lotes_df

                # No imprimimos aquí para evitar parpadeo en reruns
                pass

        # --- Tablas finales persistentes ---
        if 'R_Nivel_Bajo' in state:
            st.markdown("#### R_Nivel_Bajo")
            st.dataframe(state['R_Nivel_Bajo'], use_container_width=True)
        if 'R_Nivel_Alto' in state:
            st.markdown("#### R_Nivel_Alto")
            st.dataframe(state['R_Nivel_Alto'], use_container_width=True)






        

# =============================================================
# 6. Módulo Rastrero Out
# =============================================================
def calc_pasillo(u: str) -> str:
    if not isinstance(u, str) or len(u) < 11:
        return 'Libre'
    if u[3:5] == 'MR':
        return 'Pasillo_1'
    tramo = u[8:11]
    if tramo in ('C06','C07','C08'):
        return 'Pasillo_1'
    if tramo in ('C09','C10'):
        return 'Pasillo_2'
    if tramo in ('C11','C12'):
        return 'Pasillo_3'
    return 'Libre'


def calc_nivel(u_out: str) -> str:
    if not isinstance(u_out, str) or pd.isna(u_out):
        return ''
    if len(u_out) >= 6 and u_out[4:6] == 'MR':
        return 'B'
    last = u_out.strip()[-1]
    return 'B' if last.isdigit() and int(last) <= 2 else 'A'


def rastrero_out():
    st.subheader("📤 Rastrero Out")
    state = st.session_state.state_out

    # 1) Carga de archivos
    col1, col2, col3 = st.columns(3)
    with col1:
        asig_file = st.file_uploader("📑 Asignación (xlsx)", type="xlsx", key="asig_out")
    with col2:
        stock_file = st.file_uploader("📦 Stock (xlsx)", type="xlsx", key="stock_out")
    with col3:
        tmpl_file = st.file_uploader("🖋️ Plantilla (xlsx)", type="xlsx", key="tmpl_out")

    # 2) Lectura inicial de asignación
    headers = ['Estado','Nro. Picking','Usuario Picking','Cliente','Ubicacion',
               'Cod. Articulo','Articulo','Cant. Pick. UMS','Huella']
    if asig_file and 'df_asig_raw' not in state:
        update_status("Leyendo asignación…", 10)
        df_raw = norm_cols(pd.read_excel(asig_file))
        if not all(h in df_raw.columns for h in headers):
            update_status("⚠️ Cabeceras incorrectas", 0, False)
            return
        df = df_raw[headers].copy()
        df['Factor'] = factor(df['Huella'])
        df['Cajas_x'] = df['Cant. Pick. UMS'] / df['Factor']
        df['Concat1'] = df['Ubicacion'] + df['Cod. Articulo']
        df['Cliente_ext'] = df['Cliente'].str.split('|').str[1].fillna(df['Cliente'])
        state['df_asig_raw'] = df
        st.session_state.picks_sel = sorted(df['Nro. Picking'].dropna().unique())

        # 3) Filtro de Pickings y Fecha lado a lado
    if 'df_asig_raw' in state:
        all_picks = sorted(state['df_asig_raw']['Nro. Picking'].dropna().unique())
        # inicializar picks_sel en session_state una sola vez
        if 'picks_sel' not in st.session_state:
            st.session_state.picks_sel = all_picks
        col_date, col_filter = st.columns([1,2])
        with col_date:
            fecha = st.date_input(
                "📅 Fecha del reporte",
                st.session_state.get('fecha_out', dt.date.today()),
                key="fecha_out"
            )
        with col_filter:
            # multiselect sin default, ligando directamente a session_state
            st.multiselect(
                "🔎 Filtrar Pickings",
                options=all_picks,
                key="picks_sel"
            )

    # 4) Lectura de stock
    if stock_file and 'df_stock' not in state:
        update_status("Leyendo stock…", 10)
        df2 = norm_cols(pd.read_excel(stock_file))
        df2 = df2[['Ubicacion','Cod. Articulo','Cant. Final UMS','Huella']]
        df2['Factor'] = factor(df2['Huella'])
        df2['Cajas_y'] = df2['Cant. Final UMS'] / df2['Factor']
        df2['Concat2'] = df2['Ubicacion'] + df2['Cod. Articulo']
        state['df_stock'] = df2.groupby('Concat2', as_index=False).agg({
            'Cant. Final UMS':'sum','Cajas_y':'sum',
            'Ubicacion':'first','Cod. Articulo':'first'
        })
        update_status("Stock listo ✓", 60)

    # 5) Resúmenes automáticos tras multiselect
    picks = st.session_state.get('picks_sel', [])
    if 'df_asig_raw' in state and picks:
        df_f = state['df_asig_raw'][state['df_asig_raw']['Nro. Picking'].isin(picks)]
        state['tpick'] = df_f.groupby('Nro. Picking', as_index=False).agg({'Cant. Pick. UMS':'sum','Cajas_x':'sum'})
        state['tcli'] = df_f.groupby(['Nro. Picking','Cliente_ext'], as_index=False).agg({'Cant. Pick. UMS':'sum','Cajas_x':'sum'}).rename(columns={'Cliente_ext':'Cliente'})
        state['asign'] = df_f.groupby(['Concat1','Ubicacion','Cod. Articulo'], as_index=False)['Cajas_x'].sum()
        st.markdown("**T_Picking**"); st.dataframe(state['tpick'], use_container_width=True)
        st.markdown("**T_Clientes**"); st.dataframe(state['tcli'], use_container_width=True)
        update_status("Resúmenes listos ✓", 80)

    # 6) Generar rastrero Out
    if st.button("Generar Rastrero Out", disabled=not('asign' in state and 'df_stock' in state)):
        bd = pd.merge(state['asign'], state['df_stock'], left_on='Concat1', right_on='Concat2', how='left', suffixes=('_asig','_stk'))
        bd['UM']='CAJ'; bd['Salidas']=bd['Cajas_x']
        bd['Stock Final']=bd['Cajas_y'].fillna(0); bd['Stock Inicial']=bd['Salidas']+bd['Stock Final']
        bd['Check']=''; bd['Observacion']=''
        bd['Ubicacion_out']=bd['Ubicacion_stk'].combine_first(bd['Ubicacion_asig'])
        bd['Pasillo']=bd['Ubicacion_out'].apply(calc_pasillo)
        bd['Nivel']=bd['Ubicacion_out'].apply(calc_nivel)
        bd['Zona']=bd['Pasillo']+'_'+bd['Nivel']
        bd=bd[bd['Pasillo']!='Libre'].reset_index(drop=True)
        bd=bd.sort_values('Ubicacion_out').reset_index(drop=True)
        cols_export=['Ubicacion_out','Cod. Articulo_asig','UM','Stock Inicial','Salidas','Stock Final','Check','Observacion']
        tablas={}
        for zona, df_tab in bd.groupby('Zona'):
            if zona.startswith(('Pasillo_1','Pasillo_2','Pasillo_3')):
                df_tab=df_tab[cols_export].reset_index(drop=True)
                tablas[zona]=df_tab
                st.markdown(f"### {zona.replace('_',' — ')}"); st.dataframe(df_tab, use_container_width=True)
        state['ras_out']=tablas; update_status("Rastrero Out listo ✓", 100)


    # 7) Descarga a Excel
    if tmpl_file and 'ras_out' in state:
        tmpl_bytes = tmpl_file.read()
        try:
            wb = openpyxl.load_workbook(BytesIO(tmpl_bytes))
        except Exception as e:
            st.error(f"Error al leer la plantilla: {e}")
            return

        faltantes = [z for z in state['ras_out'] if z not in wb.sheetnames]
        if faltantes:
            st.error(f"No se encontraron las hojas: {', '.join(faltantes)} en la plantilla.")
            return

        def make_xlsx_out():
            wb2 = openpyxl.load_workbook(BytesIO(tmpl_bytes))
            for hoja, dfp in state['ras_out'].items():
                ws = wb2[hoja]
                for i, row in enumerate(dfp.itertuples(index=False), start=13):
                    for j, val in enumerate(row, start=2):
                        ws.cell(row=i, column=j, value=val)
                ws['I1'] = fecha.strftime('%d/%m/%Y')
                for idx, pk in enumerate(state['tpick']['Nro. Picking'], start=1):
                    ws.cell(row=idx, column=12, value=pk)
                last_row = 12 + len(dfp)
                if ws.max_row > last_row:
                    ws.delete_rows(last_row+1, ws.max_row - last_row)
                ws.print_area = f"B1:I{last_row}"
            buf = BytesIO(); wb2.save(buf); buf.seek(0); return buf

        fname = f"FORMATO_RASTRERO_SALIDAS_{fecha.strftime('%d.%m.%Y')}.xlsx"
        st.download_button(
            "📥 Descargar Excel Salidas",
            data=make_xlsx_out(),
            file_name=fname,
            mime="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet"
        )


# -------------------------------------------------------------
# 7. Navegación principal con streamlit-option-menu
# -------------------------------------------------------------
with st.sidebar:
        # Inyectar CSS para centrar verticalmente el nav
    st.markdown(
        """
        <style>
        /* Centra cualquier nav (incluido option_menu) verticalmente */
        [data-testid="stSidebarNav"] {
            margin-top: auto;
            margin-bottom: auto;
        }
        /* Asegura que el sidebar ocupe toda la altura */
        [data-testid="stSidebar"] > div:first-child {
            display: flex;
            flex-direction: column;
            height: 200vh;
        }
        </style>
        """,
        unsafe_allow_html=True
    )
    # Logo centrado
    st.markdown(
        """
        <div style="text-align: center; margin-bottom: 1rem;">
            <img src="https://enlinea.dinet.com.pe/e/images/logo-login.png" width="250">
        </div>
        """,
        unsafe_allow_html=True
    )

    # Espacio opcional
    st.markdown(" ")

    selected = option_menu(
        menu_title=None,
        options=[" Inicio", " Rastrero In", " Rastrero Out"],
        icons=["house", "inbox", "box-arrow-up"],
        menu_icon="cast",
        default_index=0,
        orientation="vertical",
        styles={
            "container": {
                "padding": "1!important"
                # quitamos justify-content para usar el flujo normal
            },
            "nav-link": {
                "font-size": "14px",
                "text-align": "left",
                "margin": "3px 0"
            },
            "nav-link-selected": {
                "background": "linear-gradient(135deg,#0d47a1 0%,#1976d2 100%)",
                "color": "white",
                "text-align": "left"
            },
        }
    )

# Lógica de enrutamiento
if selected == " Inicio":
    st.markdown("### Bienvenido\nSelecciona una opción del menú.")
    update_status("Elige una opción…", 0)
elif selected == " Rastrero In":
    rastrero_in()
elif selected == " Rastrero Out":
    rastrero_out()
